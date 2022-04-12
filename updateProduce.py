# -*- coding: utf-8 -*-
"""
Created on Sat Apr  2 12:19:49 2022

@author: simon
"""
##import openpyxl and loadworkbook 
import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')  ##workbook

sheet = wb.active ##active worksheet

PRICE_UPDATES = {'Garlic':3.04,
                 'Lemon': 1.13,
                    'Celery': 1.11}



## loop through the rows ,update prices        
for row in range(2, sheet.max_row): ##skips first row
    produceName = sheet.cell(row = row, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row = row, column = 2).value = PRICE_UPDATES[produceName]
        

          
wb.save('updatedProduceSales.xlsx')
       
         
        
    